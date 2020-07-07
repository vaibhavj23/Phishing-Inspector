from __future__ import division
from bs4 import BeautifulSoup
import urllib
import bs4
import socket
import whois
from datetime import datetime
import time
import pdb

# https://breakingcode.wordpress.com/2010/06/29/google-search-python/
# Previous package structure was modified. Import statements according to new structure added. Also code modified.
#from googlesearch import

import os
import sys
import re
import matplotlib
import pandas as pd
import numpy as np
from os.path import splitext
import ipaddress as ip
import tldextract
import whois
import datetime
from urllib.parse import urlparse
import matplotlib.pyplot as plt
import pandas as pd
import seaborn as sns
import pickle as pkl

import sklearn.ensemble as ek
from sklearn import tree, linear_model
from sklearn.feature_selection import SelectFromModel
from sklearn.externals import joblib
from sklearn.naive_bayes import GaussianNB
from sklearn.metrics import confusion_matrix
from sklearn.pipeline import make_pipeline
from sklearn import preprocessing
from sklearn import svm
from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import cross_val_score
from sklearn import metrics

try:
    saved_model = joblib.load('trainmodel.pkl')

except:
    df = pd.read_csv('dataset.csv')
    #df=df.sample(frac=1)
    df = df.sample(frac=1).reset_index(drop=True)
    #df.head()

    Suspicious_TLD=['zip','cricket','link','work','party','gq','kim','country','science','tk']
    Suspicious_Domain=['luckytime.co.kr','mattfoll.eu.interia.pl','trafficholder.com','dl.baixaki.com.br','bembed.redtube.comr','tags.expo9.exponential.com','deepspacer.com','funad.co.kr','trafficconverter.biz']

    # Method to count number of dots
    def countdots(url):  
        return url.count('.')

    # Method to count number of delimeters
    def countdelim(url):
        count = 0
        delim=[';','_','?','=','&']
        for each in url:
            if each in delim:
                count = count + 1
        
        return count

    import ipaddress as ip #works only in python 3

    def isip(uri):
        try:
            if ip.ip_address(uri):
                return 1
        except:
            return 0
        
    #method for domain registration details
    #method to check the presence of hyphens

    def isPresentHyphen(url):
        return url.count('-')
            
    #method to check the presence of @

    def isPresentAt(url):
        return url.count('@')

    def isPresentDSlash(url):
        return url.count('//')

    def countSubDir(url):
        return url.count('/')

    def get_ext(url):
        """Return the filename extension from url, or ''."""
        
        root, ext = splitext(url)
        return ext

    def countSubDomain(subdomain):
        if not subdomain:
            return 0
        else:
            return len(subdomain.split('.'))
        
    def countQueries(query):
        if not query:
            return 0
        else:
            return len(query.split('&'))
    def shortening_service(url):
        match = re.search('bit\.ly|goo\.gl|shorte\.st|go2l\.ink|x\.co|ow\.ly|t\.co|tinyurl|tr\.im|is\.gd|cli\.gs|'
                          'yfrog\.com|migre\.me|ff\.im|tiny\.cc|url4\.eu|twit\.ac|su\.pr|twurl\.nl|snipurl\.com|'
                          'short\.to|BudURL\.com|ping\.fm|post\.ly|Just\.as|bkite\.com|snipr\.com|fic\.kr|loopt\.us|'
                          'doiop\.com|short\.ie|kl\.am|wp\.me|rubyurl\.com|om\.ly|to\.ly|bit\.do|t\.co|lnkd\.in|'
                          'db\.tt|qr\.ae|adf\.ly|goo\.gl|bitly\.com|cur\.lv|tinyurl\.com|ow\.ly|bit\.ly|ity\.im|'
                          'q\.gs|is\.gd|po\.st|bc\.vc|twitthis\.com|u\.to|j\.mp|buzurl\.com|cutt\.us|u\.bb|yourls\.org|'
                          'x\.co|prettylinkpro\.com|scrnch\.me|filoops\.info|vzturl\.com|qr\.net|1url\.com|tweez\.me|v\.gd|'
                          'tr\.im|link\.zip\.net',
                          url)
        if match:
            return 1
        else:
            return 0

        

    featureSet = pd.DataFrame(columns=('url','no of dots','presence of hyphen','len of url','presence of at',\
    'presence of double slash','no of subdir','no of subdomain','len of domain','no of queries','is IP','presence of Suspicious_TLD',\
    'presence of suspicious domain','shortening_service','label'))

    from urllib.parse import urlparse
    import tldextract
    def getFeatures(url, label): 
        result = []
        url = str(url)
        
        #add the url to feature set
        result.append(url)
        
        #parse the URL and extract the domain information
        path = urlparse(url)
        ext = tldextract.extract(url)
        
        #counting number of dots in subdomain    
        result.append(countdots(ext.subdomain))
        
        #checking hyphen in domain   
        result.append(isPresentHyphen(path.netloc))
        
        #length of URL    
        result.append(len(url))
        
        #checking @ in the url    
        result.append(isPresentAt(path.netloc))
        
        #checking presence of double slash    
        result.append(isPresentDSlash(path.path))
        
        #Count number of subdir    
        result.append(countSubDir(path.path))
        
        #number of sub domain    
        result.append(countSubDomain(ext.subdomain))
        
        #length of domain name    
        result.append(len(path.netloc))
        
        #count number of queries    
        result.append(len(path.query))
        
        #Adding domain information
        
        #if IP address is being used as a URL     
        result.append(isip(ext.domain))
        
        #presence of Suspicious_TLD
        result.append(1 if ext.suffix in Suspicious_TLD else 0)
        
        #presence of suspicious domain
        result.append(1 if '.'.join(ext[1:]) in Suspicious_Domain else 0 )

        #presence of shortening_service
        result.append(shortening_service(url))
         
        #result.append(get_ext(path.path))
        result.append(str(label))
        return result


    for i in range(len(df)):
        features = getFeatures(df["URL"].loc[i], df["Lable"].loc[i])    
        featureSet.loc[i] = features      


    X = featureSet.drop(['url','label'],axis=1).values
    y = featureSet['label'].values


    model = {"RandomForest":ek.RandomForestClassifier(n_estimators=50)}

    from sklearn.model_selection import train_test_split
    X_train, X_test, y_train, y_test = train_test_split(X, y ,test_size=0.2)

    results = {}
    for algo in model:
        clf = model[algo]
        clf.fit(X_train,y_train)
        score = clf.score(X_test,y_test)
        print ("%s : %s " %(algo, score))
        results[algo] = score


    clf=tree.DecisionTreeClassifier(max_depth=10)
    clf.fit(X_train,y_train)
    joblib.dump(clf, 'trainmodel.pkl')
    saved_model = joblib.load('trainmodel.pkl')
    print('Model trained')

##################################################################################
import sys 
import datetime
import ssl
#from openssl import SSL
#from socket import socket
#import crypto


################-----SSL Certificate------############################## not perfect

def _get_cert_from_endpoint(server, port=443):
    try:
        cert = ssl.get_server_certificate((server, port))
    except Exception:
        #log.error('Unable to retrieve certificate from {0}'.format(server))
        cert=None
    if cert==None:
        return 0
    else:
        return 1
#print(_get_cert_from_endpoint("www.google.com"))

def start_certificate(url):
    status = []

    hostname = url
    h = [(x.start(0), x.end(0)) for x in re.finditer('https://|http://|www.|https://www.|http://www.', hostname)]
    #print(h)
    z = int(len(h))
    #print(z)
    if z != 0:
        y = h[0][1]
        hostname = hostname[y:]
        h = [(x.start(0), x.end(0)) for x in re.finditer('/', hostname)]
        z = int(len(h))
        if z != 0:
            hostname = hostname[:h[0][0]]
    
    cert_result=_get_cert_from_endpoint(hostname)
    return cert_result


################-----Domain Information------##############################

from bs4 import BeautifulSoup
import urllib
import bs4
import re
import socket
import whois
from datetime import datetime
import time
import pdb

# https://breakingcode.wordpress.com/2010/06/29/google-search-python/
# Previous package structure was modified. Import statements according to new structure added. Also code modified.
#from googlesearch import search

# This import is needed only when you run this file in isolation.
import sys




def domain_registration_length(domain):
    expiration_date = domain.expiration_date
    today = time.strftime('%Y-%m-%d')
    today = datetime.strptime(today, '%Y-%m-%d')

    registration_length = 0
    # Some domains do not have expiration dates. The application should not raise an error if this is the case.
    if expiration_date:
        try:
            registration_length = abs((expiration_date - today).days)
            creation_date_length=abs((today-domain.creation_date).days)
        except:
            return 0
    if registration_length / 365 <= 1 or creation_date_length / 365 <= 1:
        return 1
    else:
        return 0

def age_of_domain(domain):
    creation_date = domain.creation_date
    expiration_date = domain.expiration_date
    ageofdomain = 0
    if expiration_date:
        try:
            ageofdomain = abs((expiration_date - creation_date).days)
        except:
            return 0
    if ageofdomain / 30 < 18:
        return 1
    else:
        return 0
def start(url):
    status = []

    hostname = url
    h = [(x.start(0), x.end(0)) for x in re.finditer('https://|http://|www.|https://www.|http://www.', hostname)]
    #print(h)
    z = int(len(h))
    #print(z)
    if z != 0:
        y = h[0][1]
        hostname = hostname[y:]
        h = [(x.start(0), x.end(0)) for x in re.finditer('/', hostname)]
        z = int(len(h))
        if z != 0:
            hostname = hostname[:h[0][0]]

    #print("Hostname is - " + hostname)
    dns=1
    try:
        domain = whois.whois(hostname)
    except:
        dns = -1

    if dns == -1:
        status.append(-1)
    else:
        status.append(domain_registration_length(domain))

    if dns == -1:
        status.append(-1)
    else:
        status.append(age_of_domain(domain))
    if 1 in status:
        return 1
    else:
        return 0
    
###############################################################################

Suspicious_TLD=['zip','cricket','link','work','party','gq','kim','country','science','tk']
Suspicious_Domain=['luckytime.co.kr','mattfoll.eu.interia.pl','trafficholder.com','dl.baixaki.com.br','bembed.redtube.comr','tags.expo9.exponential.com','deepspacer.com','funad.co.kr','trafficconverter.biz']

# Method to count number of dots
def countdots(url):  
    return url.count('.')

# Method to count number of delimeters
def countdelim(url):
    count = 0
    delim=[';','_','?','=','&']
    for each in url:
        if each in delim:
            count = count + 1
    
    return count

import ipaddress as ip #works only in python 3

def isip(uri):
    try:
        if ip.ip_address(uri):
            return 1
    except:
        return 0
    
#method for domain registration details
#method to check the presence of hyphens

def isPresentHyphen(url):
    return url.count('-')
        
#method to check the presence of @

def isPresentAt(url):
    return url.count('@')

def isPresentDSlash(url):
    return url.count('//')

def countSubDir(url):
    return url.count('/')

def get_ext(url):
    """Return the filename extension from url, or ''."""
    
    root, ext = splitext(url)
    return ext

def countSubDomain(subdomain):
    if not subdomain:
        return 0
    else:
        return len(subdomain.split('.'))
    
def countQueries(query):
    if not query:
        return 0
    else:
        return len(query.split('&'))
def shortening_service(url):
    match = re.search('bit\.ly|goo\.gl|shorte\.st|go2l\.ink|x\.co|ow\.ly|t\.co|tinyurl|tr\.im|is\.gd|cli\.gs|'
                      'yfrog\.com|migre\.me|ff\.im|tiny\.cc|url4\.eu|twit\.ac|su\.pr|twurl\.nl|snipurl\.com|'
                      'short\.to|BudURL\.com|ping\.fm|post\.ly|Just\.as|bkite\.com|snipr\.com|fic\.kr|loopt\.us|'
                      'doiop\.com|short\.ie|kl\.am|wp\.me|rubyurl\.com|om\.ly|to\.ly|bit\.do|t\.co|lnkd\.in|'
                      'db\.tt|qr\.ae|adf\.ly|goo\.gl|bitly\.com|cur\.lv|tinyurl\.com|ow\.ly|bit\.ly|ity\.im|'
                      'q\.gs|is\.gd|po\.st|bc\.vc|twitthis\.com|u\.to|j\.mp|buzurl\.com|cutt\.us|u\.bb|yourls\.org|'
                      'x\.co|prettylinkpro\.com|scrnch\.me|filoops\.info|vzturl\.com|qr\.net|1url\.com|tweez\.me|v\.gd|'
                      'tr\.im|link\.zip\.net',
                      url)
    if match:
        return 1
    else:
        return 0

    

featureSet = pd.DataFrame(columns=('url','no of dots','presence of hyphen','len of url','presence of at',\
'presence of double slash','no of subdir','no of subdomain','len of domain','no of queries','is IP','presence of Suspicious_TLD',\
'presence of suspicious domain','shortening_service','label'))

from urllib.parse import urlparse
import tldextract
def getFeatures(url, label): 
    result = []
    url = str(url)
    
    #add the url to feature set
    result.append(url)
    
    #parse the URL and extract the domain information
    path = urlparse(url)
    ext = tldextract.extract(url)
    
    #counting number of dots in subdomain    
    result.append(countdots(ext.subdomain))
    
    #checking hyphen in domain   
    result.append(isPresentHyphen(path.netloc))
    
    #length of URL    
    result.append(len(url))
    
    #checking @ in the url    
    result.append(isPresentAt(path.netloc))
    
    #checking presence of double slash    
    result.append(isPresentDSlash(path.path))
    
    #Count number of subdir    
    result.append(countSubDir(path.path))
    
    #number of sub domain    
    result.append(countSubDomain(ext.subdomain))
    
    #length of domain name    
    result.append(len(path.netloc))
    
    #count number of queries    
    result.append(len(path.query))
    
    #Adding domain information
    
    #if IP address is being used as a URL     
    result.append(isip(ext.domain))
    
    #presence of Suspicious_TLD
    result.append(1 if ext.suffix in Suspicious_TLD else 0)
    
    #presence of suspicious domain
    result.append(1 if '.'.join(ext[1:]) in Suspicious_Domain else 0 )

    #presence of shortening_service
    result.append(shortening_service(url))
     
    #result.append(get_ext(path.path))
    result.append(str(label))
    return result

##################################################

##################----Add url to new dataset---########


# import openpyxl and tkinter modules
from openpyxl import *
from tkinter import *

# globally declare wb and sheet variable

# opening the existing excel file
#wb = load_workbook('urlDataset.xlsx')

# create the sheet object
#sheet = wb.active



def insert(URL,label):

    # opening the existing excel file
    wb = load_workbook('urlDataset.xlsx')

    # create the sheet object
    sheet = wb.active


    # assigning the max row and max column
    # value upto which data is written
    # in an excel sheet to the variable
    current_row = sheet.max_row
    current_column = sheet.max_column
       
    # get method returns current text
    # as string which we write into
    # excel spreadsheet at particular location
    sheet.cell(row=current_row + 1, column=1).value = URL
    sheet.cell(row=current_row + 1, column=2).value = label
       
    # save the file
    wb.save('urlDataset.xlsx')

#URL=input("URL: ")

#label=input("label: ")



############---to check website on idle(change 0 to 1 in while loop)---###########################################

while(1):
    
    site=input("Enter url : ")
    if(site=='no'):
        break
    #domain_status=start(site)
    cert_status=start_certificate(site)
    #print("domain_status ",domain_status)
    print("cert_status ",cert_status)

    if cert_status==1:
        domain_status=start(site)


    urlo=site
    result = pd.DataFrame(columns=('url','no of dots','presence of hyphen','len of url','presence of at',\
    'presence of double slash','no of subdir','no of subdomain','len of domain','no of queries','is IP','presence of Suspicious_TLD',\
    'presence of suspicious domain','shortening_service','label'))

    results = getFeatures(urlo,1)
    result.loc[0] = results
    result = result.drop(['url','label'],axis=1).values
    print(saved_model.predict(result))
    ans=saved_model.predict(result)
    predict_status=int(ans[0])
    def pred_final():
        if predict_status==1 and cert_status==0:
            return 1
        elif predict_status==0 and cert_status==0:
            return 1
        elif predict_status==1 and cert_status==1 and domain_status==1:
            return 1
        elif predict_status==0 and cert_status==1 and domain_status==1:
            return -1
        elif predict_status==1 and cert_status==1 and domain_status==0:
            return -1
        elif predict_status==0 and cert_status==1 and domain_status==0:
            return 0
    print("Final ",pred_final())
    insert(urlo,pred_final())
    print("inserted")
    
    



    


def main(url):
    urlo=url
    result = pd.DataFrame(columns=('url','no of dots','presence of hyphen','len of url','presence of at',\
    'presence of double slash','no of subdir','no of subdomain','len of domain','no of queries','is IP','presence of Suspicious_TLD',\
    'presence of suspicious domain','shortening_service','label'))

    results = getFeatures(urlo,1)
    result.loc[0] = results
    result = result.drop(['url','label'],axis=1).values
    print(saved_model.predict(result))
    ans=saved_model.predict(result)
    predict_status=int(ans[0])
    cert_status=start_certificate(urlo)
    res=1
    if cert_status==1:
        domain_status=start(urlo)
        if cert_status==1 and domain_status==0:
            res=0
        elif predict_status==1 and cert_status==0:
            res=1
        elif predict_status==0 and cert_status==0:
            res=1
        elif predict_status==1 and cert_status==1 and domain_status==1:
            res=1
        elif predict_status==0 and cert_status==1 and domain_status==1:
            res=-1
        elif predict_status==1 and cert_status==1 and domain_status==0:
            res=-1
        elif predict_status==0 and cert_status==1 and domain_status==0:
            res=0


    #adding result to new dataset
    
    insert(urlo,res)
    
    return res
        
    
    
